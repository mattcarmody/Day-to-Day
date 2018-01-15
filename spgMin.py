#!/usr/bin/env python3
# spgMin.py - Script that tracks minimum spend via automated AMEX emails.

import imapclient
import openpyxl 
import pyzmail
import re

purchaseRegex = re.compile(r'''(
	\n\n 
	(.*)	# Group 2 - Store
	\n\$
	(\S+)	# Group 3 - Purchase Amount
	\*\n
	(.*)	# Group 4 - Date
	\n\n
	)''', re.VERBOSE)

conn = imapclient.IMAPClient('imap.gmail.com', ssl=True)
conn.login(input('email'), input('passwd'))

# Pull out emails.
conn.select_folder('Finance/American Express', readonly=True)
UIDs = conn.search(['SINCE', '12-Jan-2017', 'SUBJECT', 'Large Purchase Approved', 'BODY', 'Starwood Preferred Guest'])

# Loop over emails
for i in range(len(UIDs)):
    rawMessage = conn.fetch([UIDs[i]], ['BODY[]', 'FLAGS'])

	# Extract email text.
    message = pyzmail.PyzMessage.factory(rawMessage[UIDs[i]][b'BODY[]'])
    payloadText = message.text_part.get_payload().decode('UTF-8')
    
    # TEMP: Save text to .txt file and then read it back in.
    # Without doing this, regex won't find match in payloadText. Why??
    amexTextFile = open('/home/matt/amexProject/amexText.txt', 'w')
    amexTextFile.write(payloadText) 
    amexTextFile.close()
    amexTextFile2 = open('/home/matt/amexProject/amexText.txt', 'r')
    amexText = amexTextFile2.read()
    
	# Extract the purchase info and store in spreadsheet.
    mo = purchaseRegex.search(payloadText)
    if mo:
        print(mo.group(2))
        print(mo.group(3))
        print(mo.group(4))
        
        wb = openpyxl.load_workbook('spgMinData.xlsx')
        sheet = wb.active
        newLine = sheet.max_row + 1
        sheet.cell(row=newLine, column=1).value = mo.group(2)
        sheet.cell(row=newLine, column=2).value = float(mo.group(3))
        sheet.cell(row=newLine, column=3).value = mo.group(4)
        wb.save('spgMinData.xlsx')    
    else:
        print('No match.')

# Sum up spending to date.
wb = openpyxl.load_workbook('spgMinData.xlsx')
sheet = wb.active
total = 0.0
for i in range(2, sheet.max_row + 1):
    total += float(sheet['B' + str(i)].value)

# Print status of spending.
if total >= 1000:
    print("Congratulations! You met your minimum spend of $1000. "
        "You've spent $" + str(total))
else:
    print("You've spent $" + str(total) + " so far. You have until February.")

conn.logout()
